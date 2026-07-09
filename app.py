import os, uuid, secrets, warnings, base64
from datetime import datetime, timedelta
from pathlib import Path
from io import BytesIO
from functools import wraps
from email.mime.text import MIMEText

warnings.filterwarnings("ignore")

from flask import (Flask, render_template, jsonify, request,
                   redirect, url_for, session, send_file)
from werkzeug.security import check_password_hash
def generate_password_hash(pwd, method="pbkdf2:sha256"):
    from werkzeug.security import generate_password_hash as _gph
    try:
        return _gph(pwd, method=method)
    except Exception:
        return _gph(pwd, method="pbkdf2:sha256")
import pandas as pd
import sqlite3

# ── Caminhos: lê da variável de ambiente se definida, senão usa o padrão ─────
_BASE = Path(os.environ.get("FAT_BASE", Path(__file__).parent))

TMPL_DIR    = os.environ.get("FAT_TMPL",   str(_BASE / "templates"))
# Disco persistente: usa /data se existir e gravável (Render Starter), senão usa _BASE
_DATA_DIR = Path("/data")
if "FAT_DB" in os.environ:
    DB_PATH = os.environ["FAT_DB"]
elif _DATA_DIR.exists() and os.access(_DATA_DIR, os.W_OK):
    DB_PATH = str(_DATA_DIR / "faturamento.db")
else:
    DB_PATH = str(_BASE / "faturamento.db")

# Log de startup para diagnóstico no Render
print(f"[STARTUP] DB_PATH = {DB_PATH}", flush=True)
print(f"[STARTUP] /data existe = {_DATA_DIR.exists()}, gravável = {os.access(str(_DATA_DIR), os.W_OK) if _DATA_DIR.exists() else 'N/A'}", flush=True)
print(f"[STARTUP] FAT_DB env = {os.environ.get('FAT_DB', '(não definido)')}", flush=True)

CTRL_PATH   = os.environ.get("FAT_CTRL",   str(_BASE / "Controle_Medicoes.xlsx"))
CREDS_PATH  = Path(os.environ.get("FAT_CREDS",  str(_BASE / "credentials.json")))
TOKEN_ENVIO = Path(os.environ.get("FAT_TOKEN",   str(_BASE / "token_envio.json")))
FROM_EMAIL  = "energysystenfaturamento@gmail.com"
SEND_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

app = Flask(__name__, template_folder=TMPL_DIR)
app.secret_key = "energy-fat-2026"

# ── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS medicoes (
                id TEXT PRIMARY KEY,
                empresa TEXT, gestor TEXT,
                contrato_num TEXT, contrato_nome TEXT,
                obra TEXT, cod TEXT, comp TEXT,
                provisao REAL DEFAULT 0,
                medicao REAL,
                pedido TEXT, nf TEXT, venc_nf TEXT,
                retencao REAL, impostos REAL,
                status TEXT DEFAULT 'previsto',
                status_prov TEXT DEFAULT 'aberta',
                obs TEXT,
                created_at TEXT, updated_at TEXT,
                delete_requested INTEGER DEFAULT 0,
                delete_requested_by TEXT,
                delete_requested_at TEXT
            );
            CREATE TABLE IF NOT EXISTS contratos (
                num TEXT PRIMARY KEY,
                nome TEXT,
                saldo REAL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS delete_requests (
                id TEXT PRIMARY KEY,
                medicao_id TEXT,
                requested_by TEXT, requested_at TEXT,
                obra TEXT, contrato_num TEXT, contrato_nome TEXT, comp TEXT
            );
            CREATE TABLE IF NOT EXISTS folhas_recebidas (
                id TEXT PRIMARY KEY,
                n_folha TEXT UNIQUE NOT NULL,
                n_contrato TEXT,
                periodo TEXT,
                municipio TEXT,
                fornecedor TEXT,
                valor_total REAL DEFAULT 0,
                arquivo TEXT,
                data_recebimento TEXT,
                status TEXT,
                nf TEXT
            );
            CREATE TABLE IF NOT EXISTS medicao_folhas (
                id TEXT PRIMARY KEY,
                medicao_id TEXT NOT NULL,
                n_folha TEXT NOT NULL,
                valor REAL DEFAULT 0,
                periodo TEXT,
                vinculado_em TEXT,
                nf TEXT
            );
            CREATE TABLE IF NOT EXISTS usuarios (
                id TEXT PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                nome TEXT,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'engenharia',
                ativo INTEGER DEFAULT 1,
                created_at TEXT
            );
            CREATE TABLE IF NOT EXISTS sessoes_ativas (
                username TEXT PRIMARY KEY,
                nome TEXT,
                role TEXT,
                ip TEXT,
                last_seen TEXT
            );
            CREATE TABLE IF NOT EXISTS reset_tokens (
                token TEXT PRIMARY KEY,
                username TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                usado INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS edit_requests (
                id TEXT PRIMARY KEY,
                protocol TEXT,
                medicao_id TEXT,
                requested_by TEXT, requested_at TEXT,
                changes TEXT,
                status TEXT DEFAULT 'pendente',
                resolved_by TEXT, resolved_at TEXT,
                obra TEXT, contrato_num TEXT, comp TEXT
            );
            CREATE TABLE IF NOT EXISTS realocacoes (
                id TEXT PRIMARY KEY,
                medicao_id_origem TEXT NOT NULL,
                medicao_id_destino TEXT,
                comp_origem TEXT NOT NULL,
                comp_destino TEXT NOT NULL,
                valor_origem REAL DEFAULT 0,
                valor_destino REAL DEFAULT 0,
                aprovado_por TEXT,
                aprovado_em TEXT,
                obs TEXT,
                created_at TEXT
            );
        """)
        # Sempre carregar seed no startup (INSERT OR IGNORE = seguro, não sobrescreve dados novos)
        # Prioridade: /data/seed_data.json (backup dinâmico) > seed_data.json do projeto
        _data_seed = _DATA_DIR / "seed_data.json"
        seed_path = _data_seed if _data_seed.exists() else Path(os.environ.get("FAT_BASE", Path(__file__).parent)) / "seed_data.json"
        print(f"[STARTUP] Usando seed: {seed_path} (existe={seed_path.exists()})", flush=True)
        if True:
            if seed_path.exists():
                import json as _json
                with open(seed_path, encoding="utf-8") as _f:
                    _seed = _json.load(_f)
                for table, rows in _seed.items():
                    if not rows:
                        continue
                    cols = list(rows[0].keys())
                    placeholders = ",".join(["?"] * len(cols))
                    col_names = ",".join(cols)
                    for row in rows:
                        try:
                            conn.execute(
                                f"INSERT OR IGNORE INTO {table}({col_names}) VALUES({placeholders})",
                                [row.get(c) for c in cols]
                            )
                        except Exception:
                            pass

        # Seed usuários padrão se a tabela estiver vazia
        now = datetime.now().isoformat()
        if not conn.execute("SELECT 1 FROM usuarios LIMIT 1").fetchone():
            for uname, role, pwd in [
                ("admin",  "admin",      "admin123"),
                ("energy", "financeiro", "energy2026"),
            ]:
                conn.execute(
                    "INSERT OR IGNORE INTO usuarios(id,username,nome,password_hash,role,ativo,created_at) VALUES(?,?,?,?,?,1,?)",
                    (str(uuid.uuid4()), uname, uname.capitalize(),
                     generate_password_hash(pwd), role, now)
                )
        # Migrar colunas novas em edit_requests se necessário
        for col, defn in [("protocol","TEXT"), ("status","TEXT DEFAULT 'pendente'"),
                          ("resolved_by","TEXT"), ("resolved_at","TEXT")]:
            try:
                conn.execute(f"ALTER TABLE edit_requests ADD COLUMN {col} {defn}")
            except Exception:
                pass
        # Migrar campo email em usuarios
        try:
            conn.execute("ALTER TABLE usuarios ADD COLUMN email TEXT")
        except Exception:
            pass
        # Migrar coluna nf em medicao_folhas e folhas_recebidas
        try:
            conn.execute("ALTER TABLE medicao_folhas ADD COLUMN nf TEXT")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE folhas_recebidas ADD COLUMN nf TEXT")
        except Exception:
            pass
        # Migrar coluna status_prov em medicoes
        try:
            conn.execute("ALTER TABLE medicoes ADD COLUMN status_prov TEXT DEFAULT 'aberta'")
        except Exception:
            pass
        # Migrar coluna empresa em contratos
        try:
            conn.execute("ALTER TABLE contratos ADD COLUMN empresa TEXT")
        except Exception:
            pass
        # Popular empresa nos contratos a partir das medicoes
        conn.execute("""
            UPDATE contratos SET empresa = (
                SELECT empresa FROM medicoes WHERE medicoes.contrato_num = contratos.num LIMIT 1
            ) WHERE empresa IS NULL OR empresa = ''
        """)
        # Migrar pedido existente -> medicao_folhas
        now = datetime.now().isoformat()
        rows_ped = conn.execute(
            "SELECT id, pedido, medicao, comp FROM medicoes WHERE pedido IS NOT NULL AND pedido != ''"
        ).fetchall()
        for row in rows_ped:
            exists = conn.execute(
                "SELECT id FROM medicao_folhas WHERE medicao_id=? AND n_folha=?",
                (row["id"], row["pedido"])
            ).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO medicao_folhas(id,medicao_id,n_folha,valor,periodo,vinculado_em) VALUES(?,?,?,?,?,?)",
                    (str(uuid.uuid4()), row["id"], row["pedido"],
                     row["medicao"] or 0, row["comp"], now)
                )

    # Exportar snapshot para /data/seed_data.json logo após init
    # (chamado após o with, então já está commitado)
    _export_seed_to_disk()


def _read_controle():
    if not Path(CTRL_PATH).exists():
        # Render: ler do banco de dados
        try:
            with get_db() as conn:
                rows = conn.execute("SELECT * FROM folhas_recebidas ORDER BY data_recebimento DESC").fetchall()
            return [dict(r) for r in rows]
        except Exception:
            return []
    try:
        df = pd.read_excel(CTRL_PATH, sheet_name="Medições")
        df.columns = [
            "data_recebimento", "n_folha", "n_contrato",
            "periodo_inicio", "periodo_fim",
            "municipio", "fornecedor", "valor_total", "arquivo", "status",
        ]
        df = df[df["n_folha"].notna()].copy()
        df["n_contrato"] = df["n_contrato"].apply(
            lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in ('', 'nan') else ''
        ).str.strip()
        df["n_folha"]     = df["n_folha"].astype(str).str.strip()
        df["valor_total"] = pd.to_numeric(df["valor_total"], errors="coerce").fillna(0)
        df["periodo_inicio"] = pd.to_datetime(df["periodo_inicio"], format="%d/%m/%Y", errors="coerce")
        df["periodo"] = df["periodo_inicio"].dt.strftime("%Y-%m")

        # Regra: período vazio → mês anterior à data de recebimento
        df["data_recebimento_dt"] = pd.to_datetime(df["data_recebimento"], format="%d/%m/%Y", errors="coerce")
        def periodo_fallback(row):
            if str(row["periodo"]) not in ("", "nan", "NaT", "None"):
                return row["periodo"]
            if pd.notna(row["data_recebimento_dt"]):
                from dateutil.relativedelta import relativedelta
                mes_ant = row["data_recebimento_dt"] - relativedelta(months=1)
                return mes_ant.strftime("%Y-%m")
            return ""
        df["periodo"] = df.apply(periodo_fallback, axis=1)

        return [{
            "n_folha":          str(row["n_folha"]),
            "n_contrato":       str(row["n_contrato"]),
            "periodo":          str(row["periodo"] or ""),
            "municipio":        str(row.get("municipio") or ""),
            "fornecedor":       str(row.get("fornecedor") or ""),
            "valor_total":      float(row["valor_total"]),
            "arquivo":          str(row.get("arquivo") or ""),
            "data_recebimento": str(row.get("data_recebimento") or ""),
            "status":           str(row.get("status") or ""),
        } for _, row in df.iterrows()]
    except Exception:
        return []

# ── Exportar DB atual para /data/seed_data.json (backup dinâmico) ────────────

def _export_seed_to_disk():
    """Exporta o DB atual para /data/seed_data.json para sobreviver a deploys."""
    if not (_DATA_DIR.exists() and os.access(str(_DATA_DIR), os.W_OK)):
        return
    import json as _j
    tables = ["medicoes", "contratos", "folhas_recebidas", "medicao_folhas",
              "usuarios", "edit_requests", "realocacoes", "delete_requests"]
    export = {}
    try:
        with get_db() as conn:
            for t in tables:
                try:
                    rows = conn.execute(f"SELECT * FROM {t}").fetchall()
                    export[t] = [dict(r) for r in rows]
                except Exception:
                    export[t] = []
        out = _DATA_DIR / "seed_data.json"
        with open(out, "w", encoding="utf-8") as f:
            _j.dump(export, f, ensure_ascii=False, default=str)
        print(f"[BACKUP] Seed exportado para {out} ({sum(len(v) for v in export.values())} registros)", flush=True)
    except Exception as e:
        print(f"[BACKUP] Erro ao exportar seed: {e}", flush=True)


# ── Auto-restore seed se banco vazio ─────────────────────────────────────────

def _auto_seed():
    try:
        with get_db() as conn:
            if conn.execute("SELECT COUNT(*) FROM medicoes").fetchone()[0] > 0:
                return
        _run_seed()
    except Exception:
        pass

def _run_seed():
    import json as _j
    seed_path = Path(os.environ.get("FAT_BASE", Path(__file__).parent)) / "seed_data.json"
    if not seed_path.exists():
        return
    with open(seed_path, encoding="utf-8") as f:
        seed = _j.load(f)
    with get_db() as conn:
        for table, rows in seed.items():
            if not rows:
                continue
            cols = list(rows[0].keys())
            ph = ",".join(["?"] * len(cols))
            col_names = ",".join(cols)
            for row in rows:
                try:
                    conn.execute(f"INSERT OR IGNORE INTO {table}({col_names}) VALUES({ph})",
                                 [row.get(c) for c in cols])
                except Exception:
                    pass

@app.before_request
def ensure_seed():
    if request.endpoint in (None, "static"):
        return
    _auto_seed()

# ── Auth ──────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return wrapper


@app.route("/api/admin/import-jun2026-7696", methods=["POST"])
@login_required
def api_import_jun2026_7696():
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    FOLHAS = [
        ("1011974296","4600027696",3612.85,"GOIANIA"),
        ("1011974297","4600027696",6164.04,"ANAPOLIS"),
        ("1011974298","4600027696",31423.77,"ANAPOLIS"),
        ("1011974299","4600027696",3612.85,"GOIANIA"),
        ("1011974302","4600027696",3614.28,"GOIANIA"),
        ("1011974304","4600027696",3614.28,"GOIANIA"),
        ("1011974305","4600027696",3612.85,"GOIANIA"),
        ("1011974306","4600027696",3612.85,"GOIANIA"),
        ("1011974307","4600027696",3612.85,"GOIANIA"),
        ("1011974308","4600027696",3612.85,"GOIANIA"),
        ("1011974309","4600027696",3612.85,"GOIANIA"),
        ("1011974310","4600027696",5237.25,"FORMOSA"),
        ("1011974311","4600027696",3053.51,"IPORA"),
        ("1011974312","4600027696",699.60,"IPORA"),
        ("1011974313","4600027696",56344.65,"IPORA"),
        ("1011974314","4600027696",17627.97,"LUZIANIA"),
        ("1011974315","4600027696",731.81,"SAO LUIZ DE MONTES BELOS"),
        ("1011974316","4600027696",1209.81,"SAO LUIZ DE MONTES BELOS"),
        ("1011974317","4600027696",720.22,"SAO LUIZ DE MONTES BELOS"),
        ("1011974318","4600027696",957.28,"SAO LUIZ DE MONTES BELOS"),
        ("1011974320","4600027696",1504.85,"SAO LUIZ DE MONTES BELOS"),
        ("1011974321","4600027696",914.76,"SAO LUIZ DE MONTES BELOS"),
        ("1011974322","4600027696",13068.24,"SAO LUIZ DE MONTES BELOS"),
        ("1011974323","4600027696",1167.29,"SAO LUIZ DE MONTES BELOS"),
        ("1011974324","4600027696",828.44,"SAO LUIZ DE MONTES BELOS"),
        ("1011974325","4600027696",2972.34,"SAO LUIZ DE MONTES BELOS"),
        ("1011974326","4600027696",2664.41,"SAO LUIZ DE MONTES BELOS"),
        ("1011974327","4600027696",1225.27,"SAO LUIZ DE MONTES BELOS"),
        ("1011974328","4600027696",778.19,"SAO LUIZ DE MONTES BELOS"),
        ("1011974329","4600027696",1019.12,"SAO LUIZ DE MONTES BELOS"),
        ("1011974330","4600027696",3612.85,"GOIANIA"),
        ("1011974331","4600027696",2933.64,"IPORA"),
        ("1011974332","4600027696",59855.09,"SAO LUIZ DE MONTES BELOS"),
        ("1011974333","4600027696",2184.14,"SAO LUIZ DE MONTES BELOS"),
        ("1011974334","4600027696",1369.79,"IPORA"),
        ("1011974335","4600027696",2739.58,"IPORA"),
        ("1011974884","4600027696",470.27,"IPORA"),
        ("1011974885","4600027696",583.65,"SAO LUIZ DE MONTES BELOS"),
        ("1011974886","4600027696",271.85,"SAO LUIZ DE MONTES BELOS"),
        ("1011974887","4600027696",117.24,"SAO LUIZ DE MONTES BELOS"),
        ("1011974888","4600027696",1743.45,"SAO LUIZ DE MONTES BELOS"),
    ]
    inseridas = atualizadas = 0
    with get_db() as conn:
        for folha, contrato, valor, municipio in FOLHAS:
            existe = conn.execute("SELECT id FROM folhas_recebidas WHERE n_folha=?", (folha,)).fetchone()
            if existe:
                conn.execute("""UPDATE folhas_recebidas SET n_contrato=?,valor_total=?,municipio=?,
                    periodo='2026-06',data_recebimento='2026-07-08',status='Processado',
                    fornecedor='ENERGY SYSTEN SERVICOS ESPECIALIZAD' WHERE n_folha=?""",
                    (contrato, valor, municipio, folha))
                atualizadas += 1
            else:
                conn.execute("""INSERT INTO folhas_recebidas(id,n_folha,n_contrato,valor_total,municipio,
                    periodo,data_recebimento,status,fornecedor)
                    VALUES(?,?,?,?,?,'2026-06','2026-07-08','Processado','ENERGY SYSTEN SERVICOS ESPECIALIZAD')""",
                    (str(uuid.uuid4()), folha, contrato, valor, municipio))
                inseridas += 1
    return jsonify({"ok": True, "inseridas": inseridas, "atualizadas": atualizadas})

@app.route("/api/admin/fix-folhas-jul2026", methods=["POST"])
@login_required
def api_fix_folhas_jul2026():
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    FIXES = [
        ("1011809376","4600027590","ANÁPOLIS",45816.19),
        ("1011809380","4600027590","BARRO ALTO",6608.89),
        ("1011809385","4600027590","GOIANÉSIA",16727.63),
        ("1011809389","4600027590","ESTRELA DO NORTE",4734.83),
        ("1011809392","4600027590","MARA ROSA",6448.11),
        ("1011809395","4600027590","MINAÇU",3021.61),
        ("1011809398","4600027590","URUAÇU",7625.69),
        ("1011809402","4600027590","FORMOSA",2164.97),
        ("1011809405","4600027590","MAMBAÍ",1873.74),
        ("1011809410","4600027590","SÃO DOMINGOS",15420.20),
        ("1011809413","4600027590","SÍTIO D'ABADIA",1873.74),
        ("1011809416","4600027590","ÁGUAS LINDAS DE GOIÁS",10732.15),
        ("1011809421","4600027590","CRISTALINA",22771.08),
        ("1011809431","4600027590","LUZIÂNIA",22030.33),
        ("1011809443","4600027590","ARAÇU",4135.29),
        ("1011809448","4600027590","INHUMAS",21169.00),
        ("1011809455","4600027590","SENADOR CANEDO",12990.94),
        ("1011809461","4600027590","SILVÂNIA",5996.38),
        ("1011809466","4600027590","TRINDADE",2164.97),
        ("1011809474","4600027590","IPORÁ",3569.85),
        ("1011809479","4600027590","BRITÂNIA",4329.94),
        ("1011809483","4600027590","ITABERAÍ",9236.76),
        ("1011809488","4600027590","ITAPURANGA",7756.48),
        ("1011809492","4600027590","ITAGUARU",7645.19),
        ("1011809496","4600027590","MATRINCHÃ",3021.61),
        ("1011809500","4600027590","MOSSÂMEDES",7757.38),
    ]
    atualizadas = 0
    with get_db() as conn:
        for folha, contrato, municipio, valor in FIXES:
            r = conn.execute("""UPDATE folhas_recebidas SET
                n_contrato=?, municipio=?, valor_total=?,
                periodo='2026-05', data_recebimento='2026-06-29',
                status='Processado', fornecedor='ENERGY SYSTEN SERVICOS ESPECIALIZAD'
                WHERE n_folha=?""", (contrato, municipio, valor, folha))
            atualizadas += r.rowcount
    return jsonify({"ok": True, "atualizadas": atualizadas})

@app.route("/api/admin/export-seed", methods=["POST"])
@login_required
def api_export_seed():
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    _export_seed_to_disk()
    out = _DATA_DIR / "seed_data.json"
    return jsonify({"ok": True, "path": str(out), "exists": out.exists()})


@app.route("/api/debug/db")
@login_required
def api_debug_db():
    info = {
        "DB_PATH": DB_PATH,
        "db_file_exists": Path(DB_PATH).exists(),
        "db_file_size_bytes": Path(DB_PATH).stat().st_size if Path(DB_PATH).exists() else 0,
        "data_dir_exists": _DATA_DIR.exists(),
        "data_dir_writable": os.access(str(_DATA_DIR), os.W_OK) if _DATA_DIR.exists() else False,
        "FAT_DB_env": os.environ.get("FAT_DB", "(não definido)"),
    }
    try:
        with get_db() as conn:
            info["medicoes_count"] = conn.execute("SELECT COUNT(*) FROM medicoes").fetchone()[0]
            info["folhas_count"] = conn.execute("SELECT COUNT(*) FROM folhas_recebidas").fetchone()[0]
            info["usuarios_count"] = conn.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
    except Exception as e:
        info["db_error"] = str(e)
    return jsonify(info)


@app.route("/login", methods=["GET", "POST"])
def login_page():
    error = None
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        with get_db() as conn:
            row = conn.execute(
                "SELECT * FROM usuarios WHERE username=? AND ativo=1", (u,)
            ).fetchone()
        if row and check_password_hash(row["password_hash"], p):
            session["user"] = row["username"]
            session["role"] = row["role"]
            session["nome"] = row["nome"] or row["username"]
            return redirect("/")
        error = "Usuário ou senha inválidos."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    if "user" in session:
        try:
            with get_db() as conn:
                conn.execute("DELETE FROM sessoes_ativas WHERE username=?", (session["user"],))
        except Exception:
            pass
    session.clear()
    return redirect("/login")

@app.route("/")
@login_required
def index():
    return render_template("index.html",
                           user=session["user"],
                           nome=session.get("nome", session["user"]),
                           role=session.get("role", "engenharia"))

@app.route("/folhas")
@login_required
def folhas_page():
    return render_template("folhas.html", user=session["user"])

# ── API ───────────────────────────────────────────────────────────────────────

@app.route("/api/medicoes")
@login_required
def api_list():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM medicoes WHERE delete_requested=0 ORDER BY comp DESC,contrato_num,obra"
        ).fetchall()
        links = conn.execute(
            "SELECT mf.id, mf.medicao_id, mf.n_folha, mf.valor, mf.periodo, mf.vinculado_em,"
            " COALESCE(NULLIF(mf.nf,''), fr.nf, '') AS nf"
            " FROM medicao_folhas mf"
            " LEFT JOIN folhas_recebidas fr ON fr.n_folha = mf.n_folha"
            " ORDER BY mf.vinculado_em"
        ).fetchall()
    by_med = {}
    for lk in links:
        by_med.setdefault(lk["medicao_id"], []).append(dict(lk))
    result = []
    for r in rows:
        d = dict(r)
        fv = by_med.get(r["id"], [])
        d["folhas_vinculadas"] = fv
        # status_prov computado:
        #  - realocada/cancelada: vem do DB
        #  - sem folhas: aberta
        #  - medido < 10% da provisão: aberta (sem expressividade)
        #  - medido 10–90%: parcial
        #  - medido ≥ 90%: cumprida
        sp = d.get("status_prov") or "aberta"
        if sp not in ("realocada", "cancelada"):
            if not fv:
                sp = "aberta"
            else:
                vl_medido  = sum(float(lk.get("valor") or 0) for lk in fv)
                provisao   = float(d.get("provisao") or 0)
                if provisao > 0:
                    pct = vl_medido / provisao
                    if pct < 0.10:
                        sp = "aberta"
                    elif pct < 0.90:
                        sp = "parcial"
                    else:
                        sp = "cumprida"
                else:
                    sp = "cumprida"   # provisão zerada mas tem folha → cumprida
        d["status_prov"] = sp
        result.append(d)
    return jsonify(result)

@app.route("/api/medicoes", methods=["POST"])
@login_required
def api_create():
    b = request.json
    now = datetime.now().isoformat()
    id_ = str(uuid.uuid4())
    with get_db() as conn:
        conn.execute("""
            INSERT INTO medicoes(id,empresa,gestor,contrato_num,contrato_nome,
                obra,cod,comp,provisao,medicao,pedido,nf,venc_nf,
                retencao,impostos,status,obs,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (id_, b.get("empresa"), b.get("gestor"),
              b.get("contrato_num"), b.get("contrato_nome"),
              b.get("obra"), b.get("cod"), b.get("comp"),
              b.get("provisao", 0), b.get("medicao") or None,
              b.get("pedido"), b.get("nf"), b.get("venc_nf") or None,
              b.get("retencao") or None, b.get("impostos") or None,
              b.get("status", "previsto"), b.get("obs"), now, now))
    return jsonify({"id": id_}), 201

@app.route("/api/medicoes/<id>", methods=["PUT"])
@login_required
def api_update(id):
    b = request.json
    now = datetime.now().isoformat()
    with get_db() as conn:
        conn.execute("""
            UPDATE medicoes SET empresa=?,gestor=?,contrato_num=?,contrato_nome=?,
                obra=?,cod=?,comp=?,provisao=?,medicao=?,pedido=?,nf=?,venc_nf=?,
                retencao=?,impostos=?,status=?,obs=?,updated_at=?
            WHERE id=?
        """, (b.get("empresa"), b.get("gestor"),
              b.get("contrato_num"), b.get("contrato_nome"),
              b.get("obra"), b.get("cod"), b.get("comp"),
              b.get("provisao", 0), b.get("medicao") or None,
              b.get("pedido"), b.get("nf"), b.get("venc_nf") or None,
              b.get("retencao") or None, b.get("impostos") or None,
              b.get("status", "previsto"), b.get("obs"), now, id))
    return jsonify({"ok": True})

@app.route("/api/medicao-folhas", methods=["POST"])
@login_required
def api_add_folha_link():
    b = request.json
    medicao_id = b.get("medicao_id")
    n_folha    = b.get("n_folha")
    valor      = float(b.get("valor", 0) or 0)
    periodo    = b.get("periodo")
    if not medicao_id or not n_folha:
        return jsonify({"erro": "medicao_id e n_folha obrigatórios"}), 400
    now = datetime.now().isoformat()
    link_id = str(uuid.uuid4())
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM medicao_folhas WHERE medicao_id=? AND n_folha=?",
            (medicao_id, n_folha)
        ).fetchone()
        if existing:
            return jsonify({"erro": "Folha já vinculada a esta obra"}), 409
        conn.execute(
            "INSERT INTO medicao_folhas(id,medicao_id,n_folha,valor,periodo,vinculado_em) VALUES(?,?,?,?,?,?)",
            (link_id, medicao_id, n_folha, valor, periodo, now)
        )
        total = conn.execute(
            "SELECT SUM(valor) FROM medicao_folhas WHERE medicao_id=?", (medicao_id,)
        ).fetchone()[0] or 0
        conn.execute(
            "UPDATE medicoes SET medicao=?, status=CASE WHEN status='previsto' THEN 'medicao' ELSE status END, updated_at=? WHERE id=?",
            (total, now, medicao_id)
        )
    return jsonify({"id": link_id, "ok": True}), 201

@app.route("/api/medicao-folhas/<id>", methods=["DELETE"])
@login_required
def api_remove_folha_link(id):
    now = datetime.now().isoformat()
    with get_db() as conn:
        lk = conn.execute("SELECT * FROM medicao_folhas WHERE id=?", (id,)).fetchone()
        if not lk:
            return jsonify({"erro": "Não encontrado"}), 404
        medicao_id = lk["medicao_id"]
        conn.execute("DELETE FROM medicao_folhas WHERE id=?", (id,))
        total = conn.execute(
            "SELECT SUM(valor) FROM medicao_folhas WHERE medicao_id=?", (medicao_id,)
        ).fetchone()[0] or None
        # Se não restam folhas, volta a previsto
        new_status_sql = "CASE WHEN (SELECT COUNT(*) FROM medicao_folhas WHERE medicao_id=?) = 0 THEN 'previsto' ELSE status END"
        conn.execute(
            f"UPDATE medicoes SET medicao=?, status={new_status_sql}, updated_at=? WHERE id=?",
            (total, medicao_id, now, medicao_id)
        )
    return jsonify({"ok": True})

@app.route("/api/medicoes/<id>/solicitar-exclusao", methods=["POST"])
@login_required
def api_delete_request(id):
    now = datetime.now().isoformat()
    with get_db() as conn:
        row = conn.execute("SELECT * FROM medicoes WHERE id=?", (id,)).fetchone()
        if not row:
            return jsonify({"erro": "Não encontrado"}), 404
        row = dict(row)
        if session.get("role") == "admin":
            conn.execute("DELETE FROM medicoes WHERE id=?", (id,))
            return jsonify({"ok": True})
        conn.execute(
            "UPDATE medicoes SET delete_requested=1,delete_requested_by=?,delete_requested_at=? WHERE id=?",
            (session["user"], now, id))
        conn.execute("""
            INSERT INTO delete_requests(id,medicao_id,requested_by,requested_at,obra,contrato_num,contrato_nome,comp)
            VALUES(?,?,?,?,?,?,?,?)
        """, (str(uuid.uuid4()), id, session["user"], now,
              row.get("obra"), row.get("contrato_num"), row.get("contrato_nome"), row.get("comp")))
    return jsonify({"ok": True})

@app.route("/api/contratos")
@login_required
def api_contratos():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM contratos ORDER BY num").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/contratos/<num>", methods=["PUT"])
@login_required
def api_update_contrato(num):
    b = request.json
    with get_db() as conn:
        conn.execute("INSERT OR REPLACE INTO contratos(num,nome,saldo) VALUES(?,?,?)",
                     (num, b.get("nome", ""), b.get("saldo", 0)))
    return jsonify({"ok": True})

@app.route("/api/folhas/sync", methods=["POST"])
@login_required
def api_folhas_sync():
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    rows = request.json or []
    if not rows:
        return jsonify({"erro": "Nenhuma folha enviada"}), 400
    inseridas = 0
    atualizadas = 0
    with get_db() as conn:
        for row in rows:
            try:
                n_folha = row["n_folha"]
                existe = conn.execute(
                    "SELECT id FROM folhas_recebidas WHERE n_folha=?", (n_folha,)
                ).fetchone()
                if existe:
                    conn.execute(
                        """UPDATE folhas_recebidas SET
                            n_contrato=?, periodo=?, municipio=?, fornecedor=?,
                            valor_total=?, arquivo=?, data_recebimento=?, status=?, nf=?
                           WHERE n_folha=?""",
                        (row.get("n_contrato",""), row.get("periodo",""),
                         row.get("municipio",""), row.get("fornecedor",""),
                         row.get("valor_total",0), row.get("arquivo",""),
                         row.get("data_recebimento",""), row.get("status",""),
                         row.get("nf"), n_folha)
                    )
                    atualizadas += 1
                else:
                    conn.execute(
                        "INSERT INTO folhas_recebidas(id,n_folha,n_contrato,periodo,municipio,fornecedor,valor_total,arquivo,data_recebimento,status,nf) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                        (row.get("id") or str(uuid.uuid4()), n_folha, row.get("n_contrato",""),
                         row.get("periodo",""), row.get("municipio",""), row.get("fornecedor",""),
                         row.get("valor_total",0), row.get("arquivo",""), row.get("data_recebimento",""),
                         row.get("status",""), row.get("nf"))
                    )
                    inseridas += 1
            except Exception:
                pass
    return jsonify({"ok": True, "inseridas": inseridas, "atualizadas": atualizadas})

@app.route("/api/folhas")
@login_required
def api_folhas():
    import json as _json
    data = _read_controle()
    resp_str = _json.dumps(data, ensure_ascii=False, allow_nan=False)
    return app.response_class(response=resp_str, status=200, mimetype='application/json')

@app.route("/api/escanear-nfs", methods=["POST"])
@login_required
def api_escanear_nfs():
    """Escaneia pastas de NFs (FINANCEIRO e ENERGY CONSTRUÇÕES) e atualiza medicao_folhas."""
    import os, re as _re, urllib.parse

    def normalizar_pasta(path):
        """Converte smb://host/share/caminho → /Volumes/share/caminho (macOS).
        Se o caminho já for local (/Volumes/... ou outro), retorna sem alteração."""
        path = path.strip()
        if path.lower().startswith("smb://"):
            # smb://192.168.1.10/SHARE_NAME/rest/of/path
            sem_proto = path[6:]  # remove 'smb://'
            # URL-decode caso venha com %20 etc
            sem_proto = urllib.parse.unquote(sem_proto)
            # Remove o host (tudo até a primeira barra)
            slash = sem_proto.find('/')
            if slash == -1:
                return path  # sem barra → não consegue converter
            sem_host = sem_proto[slash:]  # /SHARE_NAME/rest/...
            return '/Volumes' + sem_host
        return path

    b = request.json or {}

    # Aceita lista de pastas ou pasta única
    PASTAS_PADRAO = [
        "/Volumes/FINANCEIRO/CONTAS A RECEBER/NOTAS FISCAIS EMITIDAS/2026/05 MAIO",
        "/Volumes/ENERGY CONSTRUÇÕES/05- DEPARTAMENTO FINANCEIRO/01 - CONTAS A RECEBER/01 - NOTAS FISCAIS EMITIDAS/2026/05 - MAIO",
    ]

    pastas_body = b.get("pastas", [])          # lista enviada pelo modal
    pasta_extra  = b.get("pasta", "").strip()  # campo único legado

    if pastas_body:
        PASTAS_PADRAO = [normalizar_pasta(p) for p in pastas_body if p.strip()]
    elif pasta_extra:
        PASTAS_PADRAO = [normalizar_pasta(pasta_extra)]

    # Montar mapa folha→NF varrendo todas as pastas disponíveis
    mapa = {}
    pastas_ok = []
    pastas_erro = []
    for pasta in PASTAS_PADRAO:
        if not os.path.isdir(pasta):
            # Tenta variações comuns de capitalização / sufixo de montagem
            # macOS pode montar como "/Volumes/SHARE 1", "/Volumes/SHARE-1", etc.
            encontrado = False
            base_volumes = '/Volumes'
            if pasta.startswith(base_volumes + '/'):
                rest = pasta[len(base_volumes)+1:]
                share = rest.split('/')[0]
                sub   = rest[len(share):]  # tudo depois do share name
                # Verifica variantes: SHARE, SHARE 1, SHARE-1
                for suffix in ['', ' 1', ' 2', '-1', '_1']:
                    candidato = f'{base_volumes}/{share}{suffix}{sub}'
                    if os.path.isdir(candidato):
                        pasta = candidato
                        encontrado = True
                        break
            if not encontrado:
                pastas_erro.append(pasta)
                continue
        pastas_ok.append(pasta)
        for root, dirs, files in os.walk(pasta):
            for fname in files:
                if not fname.lower().endswith('.pdf'):
                    continue
                if 'CANCELADA' in fname.upper() or fname.upper().startswith('FOLHA DE REGISTRO'):
                    continue
                # Padrão: NF_FOLHA_... (folha começa com 10 e tem 10 dígitos)
                m = _re.match(r'^(\d+)_(10\d{8,})_', fname)
                if m:
                    mapa[m.group(2)] = m.group(1)

    if not mapa and not pastas_ok:
        # Monta mensagem de diagnóstico
        vols = []
        if os.path.isdir('/Volumes'):
            vols = os.listdir('/Volumes')
        msg = "Nenhuma pasta encontrada. "
        if vols:
            msg += f"Volumes montados atualmente: {', '.join(vols)}. "
        msg += "Verifique se o compartilhamento de rede está conectado no Finder."
        return jsonify({"erro": msg, "pastas_erro": pastas_erro}), 400

    now = datetime.now().isoformat()
    atualizadas = 0
    detalhes = []
    with get_db() as conn:
        links = conn.execute(
            "SELECT mf.id, mf.n_folha, mf.medicao_id FROM medicao_folhas mf"
        ).fetchall()
        for lk in links:
            nf = mapa.get(lk["n_folha"])
            if nf:
                conn.execute("UPDATE medicao_folhas SET nf=? WHERE id=?", (nf, lk["id"]))
                conn.execute(
                    "UPDATE medicoes SET status=CASE WHEN status IN ('medicao','validado','aprovado') THEN 'nf_emitida' ELSE status END, updated_at=? WHERE id=?",
                    (now, lk["medicao_id"])
                )
                atualizadas += 1
                detalhes.append({"n_folha": lk["n_folha"], "nf": nf})

    return jsonify({
        "ok": True,
        "atualizadas": atualizadas,
        "mapa_encontrado": len(mapa),
        "pastas_ok": pastas_ok,
        "pastas_erro": pastas_erro,
        "detalhes": detalhes
    })


@app.route("/api/nfs/sync-mapa", methods=["POST"])
@login_required
def api_nfs_sync_mapa():
    """Recebe {n_folha: nf_number, ...} e atualiza medicao_folhas.nf + status das medicoes."""
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    mapa = request.json or {}
    if not mapa:
        return jsonify({"erro": "Mapa vazio"}), 400
    now = datetime.now().isoformat()
    atualizadas = 0
    detalhes = []
    conn = get_db()
    try:
        links = conn.execute(
            "SELECT mf.id, mf.n_folha, mf.medicao_id FROM medicao_folhas mf"
        ).fetchall()
        for lk in links:
            nf = mapa.get(str(lk["n_folha"]))
            if nf:
                conn.execute("UPDATE medicao_folhas SET nf=? WHERE id=?", (str(nf), lk["id"]))
                if lk["medicao_id"]:
                    conn.execute(
                        "UPDATE medicoes SET status=CASE WHEN status IN ('medicao','validado','aprovado') THEN 'nf_emitida' ELSE status END, updated_at=? WHERE id=?",
                        (now, lk["medicao_id"])
                    )
                atualizadas += 1
                detalhes.append({"n_folha": lk["n_folha"], "nf": nf})
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"erro": str(e), "tipo": type(e).__name__, "atualizadas": atualizadas}), 500
    finally:
        conn.close()
    return jsonify({"ok": True, "atualizadas": atualizadas, "detalhes": detalhes})


def _extrair_dados_nf_pdf(file_bytes):
    """Extrai NF e folha do conteúdo de uma NFS-e. Retorna (nf, folha) ou (None, None)."""
    try:
        import io, re as _re
        from pdfminer.high_level import extract_text as _pdf_extract_text
        texto = _pdf_extract_text(io.BytesIO(file_bytes))
        NF_PADROES = [
            _re.compile(r'N[uú]mero\s+da\s+Nota\s+Fiscal\s*[\n:]*\s*(\d{3,6})', _re.IGNORECASE),
            _re.compile(r'nota\s+fiscal[^\d]{0,20}(\d{3,6})', _re.IGNORECASE),
            _re.compile(r'\bNFS?-?e\b[^\d]{0,20}(\d{3,6})', _re.IGNORECASE),
            _re.compile(r'\bNF[.\s\-]*n[°º]?[.\s]*(\d{3,6})', _re.IGNORECASE),
        ]
        FOLHA_PADROES = [
            _re.compile(r'FOLHA\s*DE\s*REGISTRO\s*:?\s*(\d{8,})', _re.IGNORECASE),
            _re.compile(r'FOLHADEREGISTRO\s*:?\s*(\d{8,})', _re.IGNORECASE),
            _re.compile(r'Folha\s*:\s*(\d{8,})', _re.IGNORECASE),
            _re.compile(r'N[°º]?\s*DA\s*FOLHA\s*:?\s*(\d{8,})', _re.IGNORECASE),
        ]
        nf    = next((m.group(1) for p in NF_PADROES    for m in [p.search(texto)] if m), None)
        folha = next((m.group(1) for p in FOLHA_PADROES for m in [p.search(texto)] if m), None)
        return nf, folha
    except Exception:
        pass
    return None, None


@app.route("/api/nfs/upload", methods=["POST"])
@login_required
def api_nfs_upload():
    """Recebe PDFs de NF pelo navegador. Suporta padrões:
       1. NF_FOLHA_CTR_MUN.pdf  — extrai tudo do nome
       2. NF_DESCRICAO.pdf      — NF no nome, folha extraída do PDF
       3. qualquer nome         — NF e folha extraídas do PDF
    """
    import re as _re
    # Padrão 1 completo: 1284_1011609952_4600019416_PARA.pdf
    PADRAO_COMPLETO = _re.compile(r'^(\d+)_(\d{8,})_\d+_.+\.pdf$', _re.IGNORECASE)
    # Padrão 2: 1407_1011638406_EQUATORIAL PARA.pdf  (NF + folha no nome, sem contrato)
    PADRAO_NF_FOLHA = _re.compile(r'^(\d{3,6})_(\d{8,})_.+\.pdf$', _re.IGNORECASE)
    # Padrão 3: 1284_EQUATORIAL_PARA.pdf  (só NF no nome, folha vem do PDF)
    PADRAO_NF_NOME  = _re.compile(r'^(\d{3,6})_[^0-9].+\.pdf$', _re.IGNORECASE)
    arquivos = request.files.getlist("files")
    if not arquivos:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400
    now = datetime.now().isoformat()
    resultado = []
    conn = get_db()
    try:
        folhas_db = {r["n_folha"]: dict(r) for r in conn.execute("SELECT * FROM folhas_recebidas").fetchall()}
        for arq in arquivos:
            nome = arq.filename
            file_bytes = arq.read()
            m_completo = PADRAO_COMPLETO.match(nome) or PADRAO_NF_FOLHA.match(nome)
            if m_completo:
                num_nf  = m_completo.group(1)
                n_folha = m_completo.group(2)
            else:
                # Tentar extrair NF e folha do conteúdo do PDF
                nf_pdf, folha_pdf = _extrair_dados_nf_pdf(file_bytes)
                # NF: preferir do nome se seguir padrão NF_DESCRICAO
                m_nf = PADRAO_NF_NOME.match(nome)
                num_nf  = m_nf.group(1) if m_nf else nf_pdf
                n_folha = folha_pdf
                if not num_nf:
                    resultado.append({"arquivo": nome, "status": "nf_nao_encontrada",
                                      "msg": "Não foi possível identificar o número da NF"})
                    continue
                if not n_folha:
                    resultado.append({"arquivo": nome, "nf": num_nf, "status": "nf_sem_folha",
                                      "msg": f"NF {num_nf} identificada — informe o número da folha para vincular"})
                    continue
            if n_folha in folhas_db:
                f = folhas_db[n_folha]
                nf_atual = f.get("nf") or ""
                if nf_atual == num_nf:
                    resultado.append({"arquivo": nome, "nf": num_nf, "n_folha": n_folha, "status": "ja_vinculada"})
                    continue
                conn.execute("UPDATE folhas_recebidas SET nf=? WHERE n_folha=?", (num_nf, n_folha))
                links = conn.execute("SELECT id, medicao_id FROM medicao_folhas WHERE n_folha=?", (n_folha,)).fetchall()
                for lk in links:
                    conn.execute("UPDATE medicao_folhas SET nf=? WHERE id=?", (num_nf, lk["id"]))
                    conn.execute(
                        "UPDATE medicoes SET status=CASE WHEN status IN ('medicao','validado','aprovado') THEN 'nf_emitida' ELSE status END, updated_at=? WHERE id=?",
                        (now, lk["medicao_id"])
                    )
                resultado.append({"arquivo": nome, "nf": num_nf, "n_folha": n_folha,
                                   "status": "ok", "obra": f.get("municipio",""), "links": len(links)})
            else:
                resultado.append({"arquivo": nome, "nf": num_nf, "n_folha": n_folha,
                                   "status": "folha_nao_cadastrada",
                                   "msg": f"Folha {n_folha} não encontrada no sistema"})
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"erro": str(e)}), 500
    finally:
        conn.close()
    ok    = [r for r in resultado if r["status"] == "ok"]
    sem   = [r for r in resultado if r["status"] == "folha_nao_cadastrada"]
    javin = [r for r in resultado if r["status"] == "ja_vinculada"]
    inv   = [r for r in resultado if r["status"] in ("nf_nao_encontrada", "nf_sem_folha")]
    return jsonify({"ok": True, "total": len(resultado),
                    "vinculadas": len(ok), "sem_folha": len(sem),
                    "ja_vinculadas": len(javin), "nome_invalido": len(inv),
                    "detalhes": resultado})


@app.route("/api/nfs/vincular-manual", methods=["POST"])
@login_required
def api_nfs_vincular_manual():
    """Vincula manualmente uma NF a uma folha específica."""
    data = request.json or {}
    num_nf  = str(data.get("nf", "")).strip()
    n_folha = str(data.get("n_folha", "")).strip()
    if not num_nf or not n_folha:
        return jsonify({"erro": "nf e n_folha são obrigatórios"}), 400
    now = datetime.now().isoformat()
    with get_db() as conn:
        existe = conn.execute("SELECT id FROM folhas_recebidas WHERE n_folha=?", (n_folha,)).fetchone()
        if not existe:
            return jsonify({"erro": f"Folha {n_folha} não encontrada"}), 404
        conn.execute("UPDATE folhas_recebidas SET nf=? WHERE n_folha=?", (num_nf, n_folha))
        links = conn.execute("SELECT id, medicao_id FROM medicao_folhas WHERE n_folha=?", (n_folha,)).fetchall()
        for lk in links:
            conn.execute("UPDATE medicao_folhas SET nf=? WHERE id=?", (num_nf, lk["id"]))
            conn.execute(
                "UPDATE medicoes SET status=CASE WHEN status IN ('medicao','validado','aprovado') THEN 'nf_emitida' ELSE status END, updated_at=? WHERE id=?",
                (now, lk["medicao_id"])
            )
    return jsonify({"ok": True, "nf": num_nf, "n_folha": n_folha, "links_atualizados": len(links)})

@app.route("/api/diag-db")
@login_required
def api_diag_db():
    """Diagnóstico: lista colunas das tabelas principais."""
    with get_db() as conn:
        cols = {}
        for tbl in ("medicao_folhas", "medicoes", "folhas_recebidas"):
            rows = conn.execute(f"PRAGMA table_info({tbl})").fetchall()
            cols[tbl] = [r["name"] for r in rows]
    return jsonify(cols)

@app.route("/api/preencher-datas", methods=["POST"])
@login_required
def api_preencher_datas():
    """Extrai DATA/HORA dos PDFs das folhas e preenche Data Recebimento no Controle_Medicoes.xlsx."""
    import re as _re, shutil, io
    from pdfminer.high_level import extract_text as _pdf_extract_text

    XLSX = "/Users/leonardocarmo/Documents/Claude/Projects/Faturamento/Controle_Medicoes.xlsx"

    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        ws = wb["Medições"]
    except Exception as e:
        return jsonify({"erro": f"Erro ao abrir xlsx: {e}"}), 500

    # Mapear colunas pelo cabeçalho
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_data = headers.get("Data Recebimento")
    col_folha = headers.get("Nº Folha")
    col_arquivo = headers.get("Arquivo PDF")

    if not all([col_data, col_folha, col_arquivo]):
        return jsonify({"erro": f"Colunas não encontradas. Headers: {list(headers.keys())}"}), 500

    atualizadas = 0
    erros = []
    detalhes = []

    for row in range(2, ws.max_row + 1):
        val_data = ws.cell(row, col_data).value
        # Só processa se a data estiver vazia
        if val_data and str(val_data).strip() not in ('', 'None', 'nan'):
            continue

        arquivo = ws.cell(row, col_arquivo).value
        n_folha = ws.cell(row, col_folha).value

        if not arquivo or not str(arquivo).strip():
            continue

        arquivo = str(arquivo).strip()
        if not os.path.isfile(arquivo):
            erros.append({"folha": str(n_folha), "erro": "arquivo não encontrado"})
            continue

        # Extrair DATA/HORA do PDF
        data_hora = None
        try:
            with open(arquivo, 'rb') as _f:
                text = _pdf_extract_text(io.BytesIO(_f.read()))
            m = _re.search(r'DATA(?:/HORA)?[:\s]+(\d{2}/\d{2}/\d{4})', text, _re.IGNORECASE)
            if m:
                data_hora = m.group(1)
        except Exception as e:
            erros.append({"folha": str(n_folha), "erro": str(e)})
            continue

        if data_hora:
            ws.cell(row, col_data).value = data_hora
            atualizadas += 1
            detalhes.append({"folha": str(n_folha), "data": data_hora})
        else:
            # Fallback: usar data da pasta (ex: .../04.05/...)
            m_pasta = _re.search(r'/(\d{2})\.(\d{2})/', arquivo)
            if m_pasta:
                data_hora = f"{m_pasta.group(1)}/{m_pasta.group(2)}/2026"
                ws.cell(row, col_data).value = data_hora
                atualizadas += 1
                detalhes.append({"folha": str(n_folha), "data": data_hora, "fonte": "pasta"})
            else:
                erros.append({"folha": str(n_folha), "erro": "DATA/HORA não encontrada no PDF"})

    if atualizadas > 0:
        # Backup antes de salvar
        backup = XLSX.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        shutil.copy2(XLSX, backup)
        wb.save(XLSX)

    return jsonify({
        "ok": True,
        "atualizadas": atualizadas,
        "erros": len(erros),
        "detalhes": detalhes,
        "erros_detalhes": erros[:10]
    })


@app.route("/api/edit-requests", methods=["POST"])
@login_required
def api_create_edit_request():
    b = request.json
    medicao_id = b.get("medicao_id")
    if not medicao_id:
        return jsonify({"erro": "medicao_id obrigatório"}), 400
    import json as _json
    now = datetime.now().isoformat()
    with get_db() as conn:
        row = conn.execute("SELECT * FROM medicoes WHERE id=?", (medicao_id,)).fetchone()
        if not row:
            return jsonify({"erro": "Não encontrado"}), 404
        row = dict(row)
        seq = (conn.execute("SELECT COUNT(*) FROM edit_requests").fetchone()[0] or 0) + 1
        protocol = f"ED-{datetime.now().strftime('%Y%m%d')}-{seq:04d}"
        req_id = str(uuid.uuid4())
        conn.execute("""
            INSERT INTO edit_requests(id,protocol,medicao_id,requested_by,requested_at,changes,status,obra,contrato_num,comp)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (req_id, protocol, medicao_id, session["user"], now,
              _json.dumps(b.get("changes", {})), "pendente",
              row.get("obra"), row.get("contrato_num"), row.get("comp")))
    return jsonify({"ok": True, "protocol": protocol}), 201

@app.route("/api/edit-requests")
@login_required
def api_list_edit_requests():
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM edit_requests WHERE status='pendente' ORDER BY requested_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/edit-requests/minhas")
@login_required
def api_my_edit_requests():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM edit_requests WHERE requested_by=? ORDER BY requested_at DESC LIMIT 50",
            (session["user"],)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/edit-requests/<id>/approve", methods=["POST"])
@login_required
def api_approve_edit(id):
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    import json as _json
    now = datetime.now().isoformat()
    with get_db() as conn:
        req = conn.execute("SELECT * FROM edit_requests WHERE id=?", (id,)).fetchone()
        if not req:
            return jsonify({"erro": "Não encontrado"}), 404
        b = _json.loads(req["changes"])
        conn.execute("""
            UPDATE medicoes SET empresa=?,gestor=?,contrato_num=?,contrato_nome=?,
                obra=?,cod=?,comp=?,provisao=?,medicao=?,pedido=?,nf=?,venc_nf=?,
                retencao=?,impostos=?,status=?,obs=?,updated_at=?
            WHERE id=?
        """, (b.get("empresa"), b.get("gestor"), b.get("contrato_num"), b.get("contrato_nome"),
              b.get("obra"), b.get("cod"), b.get("comp"),
              b.get("provisao", 0), b.get("medicao") or None,
              b.get("pedido"), b.get("nf"), b.get("venc_nf") or None,
              b.get("retencao") or None, b.get("impostos") or None,
              b.get("status", "previsto"), b.get("obs"), now, req["medicao_id"]))
        conn.execute(
            "UPDATE edit_requests SET status='aprovado', resolved_by=?, resolved_at=? WHERE id=?",
            (session["user"], now, id)
        )
    return jsonify({"ok": True})

@app.route("/api/edit-requests/<id>/reject", methods=["POST"])
@login_required
def api_reject_edit(id):
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    now = datetime.now().isoformat()
    with get_db() as conn:
        conn.execute(
            "UPDATE edit_requests SET status='rejeitado', resolved_by=?, resolved_at=? WHERE id=?",
            (session["user"], now, id)
        )
    return jsonify({"ok": True})

@app.route("/api/delete-requests")
@login_required
def api_delete_requests():
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM delete_requests ORDER BY requested_at DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/delete-requests/<id>/approve", methods=["POST"])
@login_required
def api_approve(id):
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    with get_db() as conn:
        req = conn.execute("SELECT * FROM delete_requests WHERE id=?", (id,)).fetchone()
        if not req:
            return jsonify({"erro": "Não encontrado"}), 404
        conn.execute("DELETE FROM medicoes WHERE id=?", (req["medicao_id"],))
        conn.execute("DELETE FROM delete_requests WHERE id=?", (id,))
    return jsonify({"ok": True})

@app.route("/api/delete-requests/<id>/reject", methods=["POST"])
@login_required
def api_reject(id):
    if session.get("role") not in ("admin", "financeiro"):
        return jsonify({"erro": "Sem permissão"}), 403
    with get_db() as conn:
        req = conn.execute("SELECT * FROM delete_requests WHERE id=?", (id,)).fetchone()
        if req:
            conn.execute("UPDATE medicoes SET delete_requested=0 WHERE id=?", (req["medicao_id"],))
        conn.execute("DELETE FROM delete_requests WHERE id=?", (id,))
    return jsonify({"ok": True})

@app.route("/api/importar/preview", methods=["POST"])
@login_required
def api_import_preview():
    file = request.files.get("arquivo")
    if not file:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400
    try:
        buf = BytesIO(file.read())
        df = pd.read_excel(buf, sheet_name="Medições", header=2)
        # Colunas: Empresa, Gestor, Nº Contrato, Nome Contrato, CC, Obra, Competência, Nº Folha, Valor Medido, NF, Obs
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {
            df.columns[0]: "empresa",
            df.columns[1]: "gestor",
            df.columns[2]: "contrato_num",
            df.columns[3]: "contrato_nome",
            df.columns[4]: "cod",
            df.columns[5]: "obra",
            df.columns[6]: "comp",
            df.columns[7]: "pedido",
            df.columns[8]: "provisao",
            df.columns[9]: "nf",
            df.columns[10]: "obs",
        } if len(df.columns) >= 11 else {}
        df = df.rename(columns=col_map)
        df = df[df["empresa"].notna() & (df["empresa"].astype(str).str.strip() != "") & (df["empresa"].astype(str).str.strip() != "nan")].copy()
        df["contrato_num"] = df["contrato_num"].apply(
            lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() not in ("", "nan") else ""
        ).str.strip()
        df["comp"] = df["comp"].apply(
            lambda x: str(x).strip() if pd.notna(x) and str(x).strip() not in ("", "nan") else ""
        )
        df["provisao"] = pd.to_numeric(df.get("provisao", pd.Series(dtype=float)), errors="coerce").fillna(0)

        import math
        def _safe(v):
            if v is None: return None
            try:
                if pd.isna(v): return None
            except: pass
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)): return None
            return v

        rows = []
        for _, r in df.iterrows():
            rows.append({
                "empresa":       _safe(r.get("empresa")),
                "gestor":        _safe(r.get("gestor")),
                "contrato_num":  _safe(r.get("contrato_num")),
                "contrato_nome": _safe(r.get("contrato_nome")),
                "cod":           _safe(r.get("cod")),
                "obra":          _safe(r.get("obra")),
                "comp":          _safe(r.get("comp")),
                "pedido":        _safe(r.get("pedido")),
                "provisao":      float(r.get("provisao", 0) or 0),
                "nf":            _safe(r.get("nf")),
                "obs":           _safe(r.get("obs")),
                "status":        "previsto",
            })

        import json as _json
        resp_str = _json.dumps({"total": len(rows), "preview": rows[:10], "rows": rows}, ensure_ascii=False, allow_nan=False)
        return app.response_class(response=resp_str, status=200, mimetype="application/json")
    except Exception as e:
        return jsonify({"erro": str(e)}), 400

@app.route("/api/importar/salvar", methods=["POST"])
@login_required
def api_import_save():
    rows = request.json.get("rows", [])
    now = datetime.now().isoformat()
    inseridas = atualizadas = 0
    with get_db() as conn:
        for r in rows:
            obra     = r.get("obra") or ""
            contrato = r.get("contrato_num") or ""
            comp     = r.get("comp") or ""
            existing = conn.execute(
                "SELECT id, status FROM medicoes WHERE obra=? AND contrato_num=? AND comp=? AND delete_requested=0",
                (obra, contrato, comp)
            ).fetchone()
            if existing:
                new_status = r.get("status", "previsto")
                if existing["status"] != "previsto":
                    new_status = existing["status"]
                conn.execute("""
                    UPDATE medicoes SET empresa=?,gestor=?,contrato_nome=?,cod=?,
                        provisao=?,obs=?,status=?,updated_at=? WHERE id=?
                """, (r.get("empresa"), r.get("gestor"), r.get("contrato_nome"),
                      r.get("cod"), r.get("provisao", 0), r.get("obs"),
                      new_status, now, existing["id"]))
                atualizadas += 1
            else:
                conn.execute("""
                    INSERT INTO medicoes(id,empresa,gestor,contrato_num,contrato_nome,
                        obra,cod,comp,provisao,obs,status,created_at,updated_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (str(uuid.uuid4()), r.get("empresa"), r.get("gestor"),
                      contrato, r.get("contrato_nome"),
                      obra, r.get("cod"), comp,
                      r.get("provisao", 0), r.get("obs"),
                      "previsto", now, now))
                inseridas += 1
        # Atualizar tabela contratos
        for r in rows:
            cnum = r.get("contrato_num")
            cnome = r.get("contrato_nome")
            emp = r.get("empresa")
            if cnum:
                conn.execute("INSERT OR IGNORE INTO contratos(num,nome,empresa,saldo) VALUES(?,?,?,0)",
                             (cnum, cnome or "", emp or ""))
    return jsonify({"inseridas": inseridas, "atualizadas": atualizadas})

@app.route("/api/provisoes-pendentes")
@login_required
def api_provisoes_pendentes():
    """Provisões de meses anteriores sem folha vinculada e não realocadas/canceladas.
    - Admin: retorna todas (visão total para acompanhamento)
    - Outros: retorna apenas as do próprio gestor (nome do usuário)
    """
    hoje = datetime.now()
    mes_atual = hoje.strftime("%Y-%m")
    # Mês anterior = exclui do modal (ainda em processamento)
    # Mostra apenas itens com 2+ meses de atraso
    from dateutil.relativedelta import relativedelta
    mes_limite = (hoje - relativedelta(months=2)).strftime("%Y-%m")
    role = session.get("role", "")
    nome = session.get("nome", session.get("user", ""))

    with get_db() as conn:
        rows = conn.execute("""
            SELECT m.*,
                   (SELECT COALESCE(SUM(mf.valor),0) FROM medicao_folhas mf WHERE mf.medicao_id = m.id) AS vl_medido,
                   (SELECT COUNT(*) FROM medicao_folhas mf WHERE mf.medicao_id = m.id) AS n_folhas
            FROM medicoes m
            WHERE m.comp <= ?
              AND m.comp != ''
              AND m.delete_requested = 0
              AND m.provisao > 0
              AND (m.status_prov IS NULL OR m.status_prov = 'aberta')
            ORDER BY m.gestor, m.comp DESC, m.contrato_num, m.obra
        """, (mes_limite,)).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        provisao  = float(d.get("provisao") or 0)
        vl_medido = float(d.get("vl_medido") or 0)
        pct = (vl_medido / provisao) if provisao > 0 else 0

        if pct < 0.10:
            d["status_prov_calc"] = "aberta"
            d["pct_medido"] = round(pct * 100, 1)
            result.append(d)
        elif pct < 0.90:
            d["status_prov_calc"] = "parcial"
            d["pct_medido"] = round(pct * 100, 1)
            result.append(d)

    return jsonify(result)

@app.route("/api/realocar", methods=["POST"])
@login_required
def api_realocar():
    """Realoca provisões não cumpridas para um mês alvo.
    - Admin/financeiro: pode realocar qualquer item.
    - Gestor: pode realocar apenas itens cujo gestor bate com seu nome.
    """
    role = session.get("role", "")
    nome = session.get("nome", session.get("user", ""))
    if role not in ("admin", "financeiro", "engenharia"):
        return jsonify({"erro": "Sem permissão."}), 403
    b = request.json or {}
    items = b.get("items", [])
    if not items:
        return jsonify({"erro": "Nenhum item"}), 400
    now = datetime.now().isoformat()
    realizados = 0
    with get_db() as conn:
        for item in items:
            mid        = item.get("medicao_id")
            valor_novo = float(item.get("valor_novo") or 0)
            comp_dest  = (item.get("comp_destino") or "").strip()
            obs        = (item.get("obs") or "").strip()
            cod_novo   = (item.get("cod") or "").strip()
            obra_nova  = (item.get("obra") or "").strip()
            if not mid or not comp_dest:
                continue
            orig = conn.execute("SELECT * FROM medicoes WHERE id=?", (mid,)).fetchone()
            if not orig:
                continue
            orig = dict(orig)
            # Gestor só pode realocar seus próprios itens
            if role not in ("admin", "financeiro"):
                gestor = (orig.get("gestor") or "").upper().strip()
                nome_u = nome.upper().strip()
                # aceita se gestor==nome completo, ou se o nome do gestor está no nome do usuário
                if gestor not in nome_u and nome_u not in gestor:
                    continue
            # Usa cod/obra do payload se informado, senão mantém do original
            cod_final  = cod_novo  or orig.get("cod")  or ""
            obra_final = obra_nova or orig.get("obra") or ""
            # Marca original como realocada
            conn.execute(
                "UPDATE medicoes SET status_prov='realocada', updated_at=? WHERE id=?",
                (now, mid)
            )
            # Cria nova provisão no mês destino
            new_id = str(uuid.uuid4())
            nova_obs = f"Realocada de {orig['comp']}." + (f" {obs}" if obs else "")
            conn.execute("""
                INSERT INTO medicoes(id,empresa,gestor,contrato_num,contrato_nome,
                    obra,cod,comp,provisao,status,status_prov,obs,created_at,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                new_id,
                orig["empresa"], orig["gestor"],
                orig["contrato_num"], orig["contrato_nome"],
                obra_final, cod_final, comp_dest,
                valor_novo, "previsto", "aberta",
                nova_obs, now, now
            ))
            # Registra no histórico de realocações
            conn.execute("""
                INSERT INTO realocacoes(id,medicao_id_origem,medicao_id_destino,
                    comp_origem,comp_destino,valor_origem,valor_destino,
                    aprovado_por,aprovado_em,obs,created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """, (
                str(uuid.uuid4()), mid, new_id,
                orig["comp"], comp_dest,
                orig["provisao"], valor_novo,
                session["user"], now, obs, now
            ))
            realizados += 1
    return jsonify({"ok": True, "realizados": realizados})

@app.route("/api/cancelar-provisao", methods=["POST"])
@login_required
def api_cancelar_provisao():
    """Marca provisões como canceladas. Gestor só pode cancelar os seus."""
    role = session.get("role", "")
    nome = session.get("nome", session.get("user", ""))
    if role not in ("admin", "financeiro", "engenharia"):
        return jsonify({"erro": "Sem permissão"}), 403
    b = request.json or {}
    ids = b.get("ids", [])
    now = datetime.now().isoformat()
    canceladas = 0
    with get_db() as conn:
        for mid in ids:
            orig = conn.execute("SELECT gestor FROM medicoes WHERE id=?", (mid,)).fetchone()
            if not orig:
                continue
            if role not in ("admin", "financeiro"):
                gestor = (orig["gestor"] or "").upper().strip()
                nome_u = nome.upper().strip()
                if gestor not in nome_u and nome_u not in gestor:
                    continue
            conn.execute(
                "UPDATE medicoes SET status_prov='cancelada', updated_at=? WHERE id=?",
                (now, mid)
            )
            canceladas += 1
    return jsonify({"ok": True, "canceladas": canceladas})

@app.route("/api/realocacoes")
@login_required
def api_historico_realocacoes():
    """Histórico de realocações."""
    with get_db() as conn:
        rows = conn.execute("""
            SELECT r.*, m.obra, m.contrato_num, m.contrato_nome, m.gestor
            FROM realocacoes r
            LEFT JOIN medicoes m ON m.id = r.medicao_id_origem
            ORDER BY r.aprovado_em DESC
        """).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/heartbeat", methods=["POST"])
@login_required
def api_heartbeat():
    now = datetime.now().isoformat()
    ip = request.remote_addr or ""
    with get_db() as conn:
        conn.execute("""
            INSERT OR REPLACE INTO sessoes_ativas(username, nome, role, ip, last_seen)
            VALUES(?,?,?,?,?)
        """, (session["user"], session.get("nome", session["user"]),
              session.get("role","engenharia"), ip, now))
    return jsonify({"ok": True})

@app.route("/api/online")
@login_required
def api_online():
    from datetime import timedelta
    cutoff = (datetime.now() - timedelta(seconds=60)).isoformat()
    with get_db() as conn:
        rows = conn.execute(
            "SELECT username, nome, role, ip, last_seen FROM sessoes_ativas WHERE last_seen >= ? ORDER BY last_seen DESC",
            (cutoff,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/usuarios")
@login_required
def api_list_usuarios():
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id,username,nome,email,role,ativo,created_at FROM usuarios ORDER BY created_at"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/usuarios", methods=["POST"])
@login_required
def api_create_usuario():
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    b = request.json
    username = (b.get("username") or "").strip()
    nome     = (b.get("nome") or "").strip()
    email    = (b.get("email") or "").strip().lower()
    pwd      = (b.get("password") or "").strip()
    role     = b.get("role", "engenharia")
    if not username or not pwd:
        return jsonify({"erro": "username e password obrigatórios"}), 400
    if role not in ("admin", "engenharia", "financeiro"):
        return jsonify({"erro": "role inválida"}), 400
    now = datetime.now().isoformat()
    uid = str(uuid.uuid4())
    try:
        with get_db() as conn:
            conn.execute(
                "INSERT INTO usuarios(id,username,nome,email,password_hash,role,ativo,created_at) VALUES(?,?,?,?,?,?,1,?)",
                (uid, username, nome or username.capitalize(), email or None, generate_password_hash(pwd), role, now)
            )
        return jsonify({"id": uid, "ok": True}), 201
    except sqlite3.IntegrityError:
        return jsonify({"erro": "Usuário já existe"}), 409

@app.route("/api/usuarios/<id>", methods=["PUT"])
@login_required
def api_update_usuario(id):
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    b = request.json
    nome  = (b.get("nome") or "").strip()
    email = (b.get("email") or "").strip().lower()
    role  = b.get("role", "engenharia")
    ativo = 1 if b.get("ativo", True) else 0
    pwd   = (b.get("password") or "").strip()
    if role not in ("admin", "engenharia", "financeiro"):
        return jsonify({"erro": "role inválida"}), 400
    with get_db() as conn:
        if pwd:
            conn.execute(
                "UPDATE usuarios SET nome=?,email=?,role=?,ativo=?,password_hash=? WHERE id=?",
                (nome, email or None, role, ativo, generate_password_hash(pwd), id)
            )
        else:
            conn.execute(
                "UPDATE usuarios SET nome=?,email=?,role=?,ativo=? WHERE id=?",
                (nome, email or None, role, ativo, id)
            )
    return jsonify({"ok": True})

@app.route("/api/usuarios/<id>", methods=["DELETE"])
@login_required
def api_delete_usuario(id):
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    with get_db() as conn:
        row = conn.execute("SELECT username FROM usuarios WHERE id=?", (id,)).fetchone()
        if not row:
            return jsonify({"erro": "Não encontrado"}), 404
        if row["username"] == session["user"]:
            return jsonify({"erro": "Não é possível desativar seu próprio usuário"}), 400
        conn.execute("UPDATE usuarios SET ativo=0 WHERE id=?", (id,))
    return jsonify({"ok": True})

@app.route("/api/carregar-seed", methods=["POST"])
@login_required
def api_carregar_seed():
    if session.get("role") != "admin":
        return jsonify({"erro": "Sem permissão"}), 403
    import json as _json
    seed_path = Path(os.environ.get("FAT_BASE", Path(__file__).parent)) / "seed_data.json"
    if not seed_path.exists():
        return jsonify({"erro": f"seed_data.json não encontrado em {seed_path}"}), 404
    with open(seed_path, encoding="utf-8") as f:
        seed = _json.load(f)
    counts = {}
    erros = []
    with get_db() as conn:
        for table, rows in seed.items():
            if not rows:
                continue
            cols = list(rows[0].keys())
            placeholders = ",".join(["?"] * len(cols))
            col_names = ",".join(cols)
            ok = 0
            for row in rows:
                try:
                    conn.execute(
                        f"INSERT OR IGNORE INTO {table}({col_names}) VALUES({placeholders})",
                        [row.get(c) for c in cols]
                    )
                    ok += 1
                except Exception as e:
                    erros.append(f"{table}: {e}")
            counts[table] = ok
    return jsonify({"ok": True, "inseridos": counts, "erros": erros[:10]})

@app.route("/api/exportar")
@login_required
def api_exportar():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM medicoes WHERE delete_requested=0").fetchall()
    df = pd.DataFrame([dict(r) for r in rows])
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    return send_file(buf, download_name="faturamento_export.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Gmail Send ────────────────────────────────────────────────────────────────

def _get_gmail_send_service():
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        creds = None
        if TOKEN_ENVIO.exists():
            creds = Credentials.from_authorized_user_file(str(TOKEN_ENVIO), SEND_SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
                TOKEN_ENVIO.write_text(creds.to_json())
            else:
                return None  # ainda não autorizado
        return build("gmail", "v1", credentials=creds)
    except Exception as e:
        print(f"[Gmail Send] Erro: {e}")
        return None

def _enviar_email_reset(dest_email, dest_nome, reset_url):
    service = _get_gmail_send_service()
    if not service:
        return False
    try:
        corpo = f"""Olá, {dest_nome}!

Recebemos uma solicitação de redefinição de senha para sua conta no Energy — Faturamento.

Clique no link abaixo para criar uma nova senha:
{reset_url}

Este link expira em 1 hora. Caso não tenha solicitado, ignore este email.

— Energy System
energysystenfaturamento@gmail.com"""
        msg = MIMEText(corpo, "plain", "utf-8")
        msg["to"]      = dest_email
        msg["from"]    = f"Energy Faturamento <{FROM_EMAIL}>"
        msg["subject"] = "Energy — Redefinição de senha"
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId="me", body={"raw": raw}).execute()
        return True
    except Exception as e:
        print(f"[Gmail Send] Erro ao enviar: {e}")
        return False

# ── Rotas de reset de senha ───────────────────────────────────────────────────

@app.route("/esqueci-senha", methods=["GET", "POST"])
def esqueci_senha():
    msg = None
    erro = None
    if request.method == "POST":
        login_or_email = request.form.get("login_or_email", "").strip().lower()
        with get_db() as conn:
            row = conn.execute(
                "SELECT * FROM usuarios WHERE (LOWER(username)=? OR LOWER(email)=?) AND ativo=1",
                (login_or_email, login_or_email)
            ).fetchone()
        if not row:
            # Mensagem genérica para não revelar se existe ou não
            msg = "Se o usuário existir, um email será enviado."
        elif not row["email"]:
            erro = "Este usuário não possui email cadastrado. Solicite ao administrador."
        else:
            token = secrets.token_urlsafe(40)
            expires = (datetime.now() + timedelta(hours=1)).isoformat()
            with get_db() as conn:
                # Invalida tokens anteriores do mesmo usuário
                conn.execute("UPDATE reset_tokens SET usado=1 WHERE username=?", (row["username"],))
                conn.execute(
                    "INSERT INTO reset_tokens(token, username, expires_at, usado) VALUES(?,?,?,0)",
                    (token, row["username"], expires)
                )
            host = request.host_url.rstrip("/")
            reset_url = f"{host}/resetar-senha/{token}"
            ok = _enviar_email_reset(row["email"], row["nome"] or row["username"], reset_url)
            if ok:
                msg = f"Email enviado para {row['email'][:3]}***. Verifique sua caixa de entrada."
            else:
                erro = "Erro ao enviar email. O serviço de email pode não estar configurado. Contate o administrador."
    return render_template("esqueci_senha.html", msg=msg, erro=erro)

@app.route("/resetar-senha/<token>", methods=["GET", "POST"])
def resetar_senha(token):
    erro = None
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM reset_tokens WHERE token=? AND usado=0", (token,)
        ).fetchone()
    if not row:
        return render_template("resetar_senha.html", invalido=True, erro=None, token=token)
    if datetime.fromisoformat(row["expires_at"]) < datetime.now():
        return render_template("resetar_senha.html", invalido=True, erro="Link expirado. Solicite um novo.", token=token)
    if request.method == "POST":
        nova = request.form.get("nova_senha", "")
        conf = request.form.get("confirmar_senha", "")
        if len(nova) < 6:
            erro = "A senha deve ter pelo menos 6 caracteres."
        elif nova != conf:
            erro = "As senhas não coincidem."
        else:
            with get_db() as conn:
                conn.execute(
                    "UPDATE usuarios SET password_hash=? WHERE username=?",
                    (generate_password_hash(nova), row["username"])
                )
                conn.execute("UPDATE reset_tokens SET usado=1 WHERE token=?", (token,))
            return render_template("resetar_senha.html", sucesso=True, invalido=False, erro=None, token=token)
    return render_template("resetar_senha.html", invalido=False, erro=erro, token=token,
                           username=row["username"])

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(debug=True, host="0.0.0.0", port=port)
