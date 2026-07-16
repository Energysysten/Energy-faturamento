"""
processar_medicoes.py
Automação de Faturamento - Energy Systen
---------------------------------------
Busca emails com folhas de medição (PDF ou XLS/XLSX) no Gmail,
salva os anexos na pasta Anexos/, extrai os dados
e atualiza a planilha Controle_Medicoes.xlsx.

Tipos de documento suportados:
  - PDF  : Folha de Registro / Folha de Medição (Equatorial Energia)
  - XLS/XLSX: Relatório Boletim / Autorização de Faturamento (Equatorial Goiás)
             → 1 arquivo pode conter N folhas (uma por linha de dados)

Dependências:
    pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client pdfplumber openpyxl

Primeira execução:
    1. Coloque o arquivo credentials.json nesta pasta
    2. Execute: python3 processar_medicoes.py
    3. Uma janela do navegador abrirá para autorização — authorize uma vez
    4. O token.json é salvo automaticamente para as próximas execuções
"""

import os
import re
import base64
import json
import datetime
import calendar
import pdfplumber
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# ── Configuração ────────────────────────────────────────────────────────────
BASE_DIR         = Path(__file__).parent
ANEXOS_DIR       = BASE_DIR / "Anexos"
PLANILHA         = BASE_DIR / "Controle_Medicoes.xlsx"
CREDENTIALS      = BASE_DIR / "credentials.json"
TOKEN            = BASE_DIR / "token.json"
LOG_FILE         = BASE_DIR / "processamento.log"
LABEL_PROCESSADO = "Medicao-Processada"

# URL da API do Render (deixe vazio para desativar o sync automático)
RENDER_API_URL  = os.environ.get("RENDER_API_URL", "https://faturamento-9pbl.onrender.com")
RENDER_API_KEY  = os.environ.get("SYNC_API_KEY", "")

# Correções de período: folhas cuja data de início na planilha difere do mês de competência.
# Adicione aqui sempre que uma folha precisar de correção manual de período.
PERIODO_OVERRIDE = {
    "1012047434": "2026-06",  # data início 26/05 mas compete em jun/2026
    "1012038827": "2026-06",  # data início 26/05 mas compete em jun/2026
    "1012038744": "2026-06",  # data início 26/05 mas compete em jun/2026
}

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/gmail.modify",
]

BORDER_COLOR = "BDD7EE"
ALT_ROW      = "D6E4F0"

# Mapa de competência abreviada → número do mês (formato dos XLS da Equatorial)
MESES_PT = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}


# ── Utilitários ─────────────────────────────────────────────────────────────
def log(msg: str):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def competencia_para_periodo(competencia: str):
    """
    Converte "ABR26" → ("01/04/2026", "30/04/2026").
    Aceita também "ABR/2026" ou "04/2026".
    Retorna ("", "") se não conseguir parsear.
    """
    comp = competencia.strip().upper()

    # Tenta formato "MMM YY" ou "MMMYY" (ex: "ABR26", "ABR 26")
    m = re.match(r"([A-Z]{3})\s*(\d{2})$", comp)
    if m:
        mes_str, ano_str = m.group(1), m.group(2)
        mes = MESES_PT.get(mes_str)
        if mes:
            ano = 2000 + int(ano_str)
            ultimo_dia = calendar.monthrange(ano, mes)[1]
            inicio = f"01/{mes:02d}/{ano}"
            fim    = f"{ultimo_dia}/{mes:02d}/{ano}"
            return inicio, fim

    # Tenta formato "MMM/YYYY" (ex: "ABR/2026")
    m = re.match(r"([A-Z]{3})/(\d{4})$", comp)
    if m:
        mes_str, ano_str = m.group(1), m.group(2)
        mes = MESES_PT.get(mes_str)
        if mes:
            ano = int(ano_str)
            ultimo_dia = calendar.monthrange(ano, mes)[1]
            inicio = f"01/{mes:02d}/{ano}"
            fim    = f"{ultimo_dia}/{mes:02d}/{ano}"
            return inicio, fim

    # Tenta formato "MM/YYYY"
    m = re.match(r"(\d{2})/(\d{4})$", comp)
    if m:
        mes, ano = int(m.group(1)), int(m.group(2))
        ultimo_dia = calendar.monthrange(ano, mes)[1]
        return f"01/{mes:02d}/{ano}", f"{ultimo_dia}/{mes:02d}/{ano}"

    return "", ""


def autenticar_gmail():
    creds = None
    if TOKEN.exists():
        creds = Credentials.from_authorized_user_file(TOKEN, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS.exists():
                raise FileNotFoundError(
                    f"\n\nArquivo credentials.json não encontrado em {CREDENTIALS}\n"
                    "Siga o guia SETUP_OAUTH.md para criar as credenciais.\n"
                )
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN, "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def obter_ou_criar_label(service, nome: str) -> str:
    """Retorna o ID de um label do Gmail, criando-o se não existir."""
    labels = service.users().labels().list(userId="me").execute().get("labels", [])
    for lb in labels:
        if lb["name"] == nome:
            return lb["id"]
    novo = service.users().labels().create(
        userId="me",
        body={"name": nome, "labelListVisibility": "labelShow", "messageListVisibility": "show"},
    ).execute()
    return novo["id"]


# ── Download de anexos ───────────────────────────────────────────────────────
def _is_pdf(filename: str, mime: str) -> bool:
    return filename.lower().endswith(".pdf") or mime == "application/pdf"


def _is_xls(filename: str, mime: str) -> bool:
    fn = filename.lower()
    return (
        fn.endswith(".xls") or fn.endswith(".xlsx")
        or mime in (
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )


def baixar_anexos(service, msg_id: str) -> dict:
    """
    Baixa todos os anexos PDF e XLS/XLSX de uma mensagem.
    Retorna: {"pdf": [Path, ...], "xls": [Path, ...]}
    """
    msg = service.users().messages().get(userId="me", id=msg_id, format="full").execute()
    resultado = {"pdf": [], "xls": []}

    def percorrer_partes(partes):
        for parte in partes:
            if "parts" in parte:
                percorrer_partes(parte["parts"])
            filename = parte.get("filename", "")
            mime     = parte.get("mimeType", "")

            if not filename:
                continue

            if _is_pdf(filename, mime):
                tipo = "pdf"
            elif _is_xls(filename, mime):
                tipo = "xls"
            else:
                continue

            att_id = parte["body"].get("attachmentId")
            if not att_id:
                continue

            att  = service.users().messages().attachments().get(
                userId="me", messageId=msg_id, id=att_id
            ).execute()
            dados = base64.urlsafe_b64decode(att["data"])

            nome_limpo = re.sub(r'[\\/*?:"<>|]', "_", filename)
            caminho = ANEXOS_DIR / nome_limpo

            if caminho.exists():
                log(f"  Anexo já existe (reprocessando): {nome_limpo}")
                resultado[tipo].append(caminho)
                continue

            caminho.write_bytes(dados)
            log(f"  Anexo salvo: {nome_limpo}")
            resultado[tipo].append(caminho)

    percorrer_partes(msg.get("payload", {}).get("parts", []))
    return resultado


# ── Extração de dados — PDF ──────────────────────────────────────────────────
def extrair_dados_pdf(caminho_pdf: Path) -> dict:
    """
    Extrai os campos da Folha de Registro usando pdfplumber + regex.
    Retorna um único dict.
    """
    texto = ""
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto += (pagina.extract_text() or "") + "\n"

    def buscar(padrao, texto, grupo=1, default=""):
        m = re.search(padrao, texto, re.IGNORECASE)
        return m.group(grupo).strip() if m else default

    folha    = buscar(r"FOLHA\s*DE\s*REGISTRO\s*:\s*(\d+)", texto)
    contrato = buscar(r"(?<![A-Z])CONTRATO\s*:\s*(\d{8,})", texto)

    periodo = re.search(
        r"PER[IÍ]ODO\s*DA\s*MEDI[ÇC][ÃA]O\s*:\s*(\d{2}[./]\d{2}[./]\d{4})\s*(\d{2}[./]\d{2}[./]\d{4})",
        texto, re.IGNORECASE
    )
    periodo_inicio = periodo.group(1).replace(".", "/") if periodo else ""
    periodo_fim    = periodo.group(2).replace(".", "/") if periodo else ""

    municipio  = buscar(r"MUNIC[IÍ]PIO\s*:\s*([A-ZÀ-Úa-zà-ú]+)", texto)
    fornecedor = buscar(r"FORNECEDOR\s*:\s*(.+?)(?:\n|$)", texto)

    total_raw = buscar(r"TOTAL\s*GERAL\s+([\d.,]+)", texto)
    try:
        valor = float(total_raw.replace(".", "").replace(",", "."))
    except ValueError:
        valor = 0.0

    return {
        "folha":          folha,
        "contrato":       contrato,
        "periodo_inicio": periodo_inicio,
        "periodo_fim":    periodo_fim,
        "municipio":      municipio.strip().title(),
        "fornecedor":     fornecedor.strip().title(),
        "valor":          valor,
        "tipo_origem":    "PDF",
    }


# ── Extração de dados — XLS/XLSX ─────────────────────────────────────────────
def extrair_dados_xls(caminho_xls: Path) -> list[dict]:
    """
    Extrai os dados do Relatório Boletim / Autorização de Faturamento (XLS/XLSX).

    Estrutura esperada (Equatorial Goiás):
      Linha 1: título  ("Boletim de Medição")
      Linha 2: subtítulo ("Autorização de Faturamento")
      Linha 3: cabeçalhos (DISTRIBUIDORA, REGIONAL, COMPETÊNCIA, TIPO MEDIÇÃO,
                           CONTRATO ANTIGO, CONTRATO, ANALISTA, PARCEIRO,
                           MUNICÍPIO, PROCESSO, FOLHA, VALOR, ...)
      Linha 4+: dados

    Retorna lista de dicts (um por linha de dados válida).
    """
    wb = openpyxl.load_workbook(caminho_xls, data_only=True)

    # Tenta localizar a aba de dados — usa a primeira aba se não encontrar "DADOS"
    sheet_name = "DADOS" if "DADOS" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Detecta linha de cabeçalho: primeira linha que contém "FOLHA", "CONTRATO"
    # ou variações do arquivo "SERVIÇOS POR FOLHA" (ex: "Nº de documento")
    MARCADORES_HEADER = {"FOLHA", "CONTRATO", "Nº DE DOCUMENTO", "NO DE DOCUMENTO", "NUMERO DE DOCUMENTO", "DOCUMENTO"}
    header_row  = None
    header_cols = {}
    for row in ws.iter_rows():
        row_vals = [str(c.value).strip().upper() if c.value else "" for c in row]
        if any(m in row_vals for m in MARCADORES_HEADER):
            header_row = row[0].row
            for idx, val in enumerate(row_vals):
                header_cols[val] = idx  # mapeia nome → índice 0-based
            break

    if header_row is None:
        log(f"  AVISO: cabeçalho não encontrado em {caminho_xls.name}. Arquivo ignorado.")
        return []

    # Campos que precisamos (tolerante a variações de nome)
    def col(nome_principal, *alternativas):
        for nome in (nome_principal, *alternativas):
            if nome in header_cols:
                return header_cols[nome]
        return None

    idx_folha      = col("FOLHA", "FOLHA DE MEDIÇÃO", "NR FOLHA", "Nº DE DOCUMENTO", "NO DE DOCUMENTO", "NUMERO DE DOCUMENTO", "DOCUMENTO")
    idx_contrato   = col("CONTRATO", "NR CONTRATO", "NÚMERO CONTRATO", "CONTRATO BÁSICO", "CONTRATO BASICO")
    idx_competenca = col("COMPETÊNCIA", "COMPETENCIA", "MÊS", "MES")
    idx_municipio  = col("MUNICÍPIO", "MUNICIPIO", "CIDADE", "LOCAL")
    idx_parceiro   = col("PARCEIRO", "FORNECEDOR", "EMPRESA")
    idx_valor      = col("VALOR", "VALOR TOTAL", "TOTAL", "VALOR LÍQUIDO", "VALOR LIQUIDO")

    registros = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Pula linhas completamente vazias
        if all(v is None for v in row):
            continue

        def get(idx):
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        folha      = get(idx_folha)
        contrato   = get(idx_contrato)
        competencia = get(idx_competenca)
        municipio  = get(idx_municipio)
        parceiro   = get(idx_parceiro)
        valor_raw  = row[idx_valor] if idx_valor is not None and idx_valor < len(row) else None

        # Ignora linhas sem folha ou contrato
        if not folha or not contrato:
            continue

        # Converte período
        periodo_inicio, periodo_fim = competencia_para_periodo(competencia)

        # Regra: período vazio → mês anterior à data de recebimento da folha
        if not periodo_inicio:
            try:
                from dateutil.relativedelta import relativedelta as _rd
                import datetime as _dt
                # data_recebimento é a data de hoje (data de processamento)
                hoje = _dt.date.today()
                mes_ant = hoje - _rd(months=1)
                ultimo_dia = calendar.monthrange(mes_ant.year, mes_ant.month)[1]
                periodo_inicio = f"01/{mes_ant.month:02d}/{mes_ant.year}"
                periodo_fim    = f"{ultimo_dia}/{mes_ant.month:02d}/{mes_ant.year}"
            except Exception:
                pass

        # Converte valor
        # Se o openpyxl já retornou um número, usa diretamente.
        # Se for string (ex: "3.612,85"), converte formato brasileiro.
        if isinstance(valor_raw, (int, float)):
            valor = float(valor_raw)
        else:
            try:
                valor = float(
                    str(valor_raw).replace("R$", "").replace(".", "").replace(",", ".").strip()
                )
            except (ValueError, TypeError):
                valor = 0.0

        registros.append({
            "folha":          folha,
            "contrato":       contrato,
            "periodo_inicio": periodo_inicio,
            "periodo_fim":    periodo_fim,
            "municipio":      municipio.strip().title(),
            "fornecedor":     parceiro.strip().title(),
            "valor":          valor,
            "tipo_origem":    "XLS",
        })

    # Agrupa por folha somando valores (um folha pode ter múltiplas linhas de serviço)
    agrupado = {}
    for r in registros:
        chave = r["folha"]
        if chave not in agrupado:
            agrupado[chave] = r.copy()
        else:
            agrupado[chave]["valor"] += r["valor"]
            # Preserva municipio/competência do primeiro registro
    return list(agrupado.values())


# ── Atualização da planilha ──────────────────────────────────────────────────
def atualizar_planilha(dados: dict, nome_arquivo: str, data_recebimento: str):
    """Acrescenta uma linha na planilha Controle_Medicoes.xlsx."""
    thin   = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.load_workbook(PLANILHA)
    ws = wb["Medições"]

    # Verifica se a folha já foi registrada
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[1]) == str(dados["folha"]):
            log(f"  Folha {dados['folha']} já está na planilha. Pulando.")
            return False

    nova_linha = ws.max_row + 1
    cor = ALT_ROW if nova_linha % 2 == 0 else "FFFFFF"

    valores = [
        data_recebimento,
        dados["folha"],
        dados["contrato"],
        dados["periodo_inicio"],
        dados["periodo_fim"],
        dados["municipio"],
        dados["fornecedor"],
        dados["valor"],
        nome_arquivo,
        "Processado",
    ]

    for col_idx, val in enumerate(valores, 1):
        cell = ws.cell(row=nova_linha, column=col_idx, value=val)
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=cor)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        if col_idx == 8:
            cell.number_format = "R$ #,##0.00"

    wb.save(PLANILHA)
    log(f"  Planilha atualizada: folha {dados['folha']} | R$ {dados['valor']:,.2f}")
    return True


def sync_folha_render(dados: dict, nome_arquivo: str, data_recebimento: str):
    """Envia a folha processada para o banco de dados do Render via API."""
    if not RENDER_API_URL or not RENDER_API_KEY:
        return

    import urllib.request

    # Converte data de DD/MM/YYYY para YYYY-MM-DD
    try:
        partes = data_recebimento.split("/")
        data_iso = f"{partes[2]}-{partes[1]}-{partes[0]}"
    except Exception:
        data_iso = str(datetime.date.today())

    # Constrói periodo no formato YYYY-MM
    periodo = ""
    try:
        if dados.get("periodo_inicio"):
            p = str(dados["periodo_inicio"])
            if "/" in p:
                parts = p.split("/")
                if len(parts) == 3:
                    periodo = f"{parts[2]}-{parts[1]}"
                elif len(parts) == 2:
                    periodo = f"{parts[1]}-{parts[0]}"
            else:
                periodo = p[:7]
    except Exception:
        pass

    payload = json.dumps([{
        "n_folha": str(dados["folha"]),
        "n_contrato": str(dados.get("contrato", "")),
        "valor_total": float(dados.get("valor", 0)),
        "municipio": str(dados.get("municipio", "")),
        "fornecedor": str(dados.get("fornecedor", "")),
        "periodo": periodo,
        "data_recebimento": data_iso,
        "arquivo": nome_arquivo,
        "status": "recebido",
    }]).encode("utf-8")

    try:
        req = urllib.request.Request(
            f"{RENDER_API_URL}/api/folhas/sync",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "X-API-Key": RENDER_API_KEY,
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            resultado = json.loads(resp.read())
            ins = resultado.get("inseridas", 0)
            atu = resultado.get("atualizadas", 0)
            log(f"  Render sync: inseridas={ins}, atualizadas={atu}")
    except Exception as e:
        log(f"  AVISO: falha ao sincronizar com Render: {e}")


# ── Fluxo principal ──────────────────────────────────────────────────────────
def main():
    ANEXOS_DIR.mkdir(exist_ok=True)
    log("=" * 60)
    log("Iniciando processamento de folhas de medição")

    service  = autenticar_gmail()
    label_id = obter_ou_criar_label(service, LABEL_PROCESSADO)

    # Busca emails com PDF ou XLS/XLSX ainda não processados
    query = (
        f"has:attachment "
        f"(filename:pdf OR filename:xls OR filename:xlsx) "
        f"-label:{LABEL_PROCESSADO}"
    )
    resultado = service.users().threads().list(userId="me", q=query, maxResults=50).execute()
    threads   = resultado.get("threads", [])

    if not threads:
        log("Nenhum email novo com PDF/XLS encontrado.")
        log(f"\nConcluído. 0 folha(s) processada(s).")
        log("=" * 60)
        return

    log(f"{len(threads)} thread(s) com anexo(s) encontrada(s).")
    processados = 0

    for thread in threads:
        thread_id   = thread["id"]
        thread_data = service.users().threads().get(
            userId="me", id=thread_id, format="metadata",
            metadataHeaders=["Date", "Subject", "From"]
        ).execute()

        for msg in thread_data.get("messages", []):
            msg_id      = msg["id"]
            headers_msg = {h["name"]: h["value"] for h in msg.get("payload", {}).get("headers", [])}
            assunto     = headers_msg.get("Subject", "Sem assunto")
            data_raw    = headers_msg.get("Date", "")
            remetente   = headers_msg.get("From", "")

            # Converte data para DD/MM/YYYY
            try:
                from email.utils import parsedate_to_datetime
                dt = parsedate_to_datetime(data_raw)
                data_recebimento = dt.strftime("%d/%m/%Y")
            except Exception:
                data_recebimento = datetime.date.today().strftime("%d/%m/%Y")

            log(f"\nProcessando: '{assunto}' de {remetente}")

            anexos = baixar_anexos(service, msg_id)
            teve_anexo = bool(anexos["pdf"] or anexos["xls"])

            if not teve_anexo:
                log("  Nenhum anexo (PDF/XLS) encontrado nesta mensagem.")
                continue

            # ── PDFs ──────────────────────────────────────────────────────────
            for pdf_path in anexos["pdf"]:
                try:
                    dados = extrair_dados_pdf(pdf_path)
                    if not dados["folha"] or not dados["contrato"]:
                        log(f"  AVISO: não foi possível extrair folha/contrato de {pdf_path.name}")
                        continue
                    if atualizar_planilha(dados, pdf_path.name, data_recebimento):
                        processados += 1
                        sync_folha_render(dados, pdf_path.name, data_recebimento)
                except Exception as e:
                    log(f"  ERRO ao processar PDF {pdf_path.name}: {e}")

            # ── XLS / XLSX ────────────────────────────────────────────────────
            for xls_path in anexos["xls"]:
                try:
                    lista = extrair_dados_xls(xls_path)
                    if not lista:
                        log(f"  AVISO: nenhum dado extraído de {xls_path.name}")
                        continue
                    log(f"  {len(lista)} folha(s) encontrada(s) em {xls_path.name}")
                    for dados in lista:
                        if atualizar_planilha(dados, xls_path.name, data_recebimento):
                            processados += 1
                            sync_folha_render(dados, xls_path.name, data_recebimento)
                except Exception as e:
                    log(f"  ERRO ao processar XLS {xls_path.name}: {e}")

            # Aplica label "Medicao-Processada" na mensagem
            service.users().messages().modify(
                userId="me", id=msg_id,
                body={"addLabelIds": [label_id]}
            ).execute()

    log(f"\nConcluído. {processados} folha(s) processada(s).")

    # Sync completo da planilha com o Render
    sync_planilha_completa_render()

    log("=" * 60)


def sync_planilha_completa_render():
    """Sincroniza todas as folhas da planilha local com o banco do Render."""
    if not RENDER_API_URL or not RENDER_API_KEY:
        return
    if not PLANILHA.exists():
        return

    import urllib.request
    import openpyxl as _opx

    try:
        wb = _opx.load_workbook(PLANILHA, read_only=True, data_only=True)
        ws = wb["Medições"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        log(f"  AVISO: não foi possível ler planilha para sync completo: {e}")
        return

    payload = []
    for row in rows:
        if not row[1]:
            continue
        try:
            n_folha    = str(row[1]).strip().split(".")[0]
            n_contrato = str(int(float(row[2]))) if row[2] and str(row[2]).strip() not in ("", "nan") else ""
            periodo    = ""
            if row[3]:
                try:
                    if hasattr(row[3], "strftime"):
                        periodo = row[3].strftime("%Y-%m")
                    else:
                        parts = str(row[3]).split("/")
                        if len(parts) == 3:
                            periodo = f"{parts[2]}-{parts[1]}"
                except Exception:
                    pass
            data_rec = ""
            if row[0]:
                try:
                    if hasattr(row[0], "strftime"):
                        data_rec = row[0].strftime("%Y-%m-%d")
                    else:
                        parts = str(row[0]).split("/")
                        if len(parts) == 3:
                            data_rec = f"{parts[2]}-{parts[1]}-{parts[0]}"
                except Exception:
                    pass
            valor = float(row[7]) if row[7] else 0.0
            # Aplica correção de período se necessário
            periodo_final = PERIODO_OVERRIDE.get(n_folha, periodo)
            payload.append({
                "n_folha": n_folha,
                "n_contrato": n_contrato,
                "valor_total": valor,
                "municipio": str(row[5] or ""),
                "fornecedor": str(row[6] or ""),
                "periodo": periodo_final,
                "data_recebimento": data_rec,
                "arquivo": str(row[8] or ""),
                "status": str(row[9] or "recebido"),
            })
        except Exception:
            continue

    if not payload:
        return

    total_ins = total_atu = 0
    for i in range(0, len(payload), 50):
        lote = payload[i:i+50]
        data = json.dumps(lote).encode("utf-8")
        try:
            req = urllib.request.Request(
                f"{RENDER_API_URL}/api/folhas/sync",
                data=data,
                headers={"Content-Type": "application/json", "X-API-Key": RENDER_API_KEY},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                r = json.loads(resp.read())
                total_ins += r.get("inseridas", 0)
                total_atu += r.get("atualizadas", 0)
        except Exception as e:
            log(f"  AVISO: falha no sync completo (lote {i//50+1}): {e}")

    log(f"  Sync completo: {total_ins} inseridas, {total_atu} atualizadas no Render")


if __name__ == "__main__":
    main()
